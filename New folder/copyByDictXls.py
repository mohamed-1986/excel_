#! Python 3
# - Copy and Paste Ranges using OpenPyXl library

import os
import openpyxl
import xlrd
#Prepare the spreadsheets to copy from and paste too.

#File to be copied
wb = xlrd.open_workbook('Sample1.xls')
copySheet = wb.sheet_by_name("Area 02")

# #File to be pasted into
template = openpyxl.load_workbook("Sample2.xlsx") #Add file name
pasteSheet = template["Area 02"] #Add Sheet name

print(copySheet.nrows)
print(copySheet.ncols )
#this returns the column place for the main headers ex. tag in column 2 and problem in column 3
def searchForWord(sheet, theWord):
    for i in range(1, 10,1):
        #Appends the row to a RowSelected list
        for j in range(1, 10,1):
            if theWord in str(sheet.cell(i,j).value).upper():
                return j
tagPaste= searchForWord(pasteSheet, "TAG")
tagCopy= searchForWord(copySheet, "TAG")

problemPaste= searchForWord(pasteSheet,"PROBLEM")
problemCopy= searchForWord(copySheet,"PROBLEM")

complainPaste= searchForWord(pasteSheet,"COMP")
complainCopy= searchForWord(copySheet,"COMP")

actionPaste= searchForWord(pasteSheet,"ACTION")
actionCopy= searchForWord(copySheet,"ACTION")

statusPaste= searchForWord(pasteSheet,"STATUS")
statusCopy= searchForWord(copySheet,"STATUS")

datePaste= searchForWord(pasteSheet,"DATE")
dateCopy= searchForWord(copySheet,"DATE")

# A dictionary for the label tags with a column number for each tag.
# later need a row iteration to loop through data.
#ex tag:2 problem:3 Complain:4
nonFilteredPasteDict= {
 "Tag": tagPaste, "Problem": problemPaste,
 "Complain": complainPaste , "Action" : actionPaste, 
 "Status": statusPaste, "Date": datePaste
 }

nonFilterdedCopyDict= {
 "Tag": tagCopy, "Problem": problemCopy,
 "Complain": complainCopy , "Action" : actionCopy, 
 "Status": statusCopy, "Date": dateCopy
 }

# drop down any key value pair if empty value
def dictValidator(d):
    new={}
    for k,v in d.items():
        if v!= None:
            new[k]=v
    return new

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(copyDict, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(12,14,1):
        #Appends the row to a RowSelected list
        rowSelected = {}
        for j in copyDict:
            if j != "Date":
                rowSelected[j] =sheet.cell(i , copyDict[j]).value
            else:
                d=sheet.cell( i , copyDict[j]).value
                #rowSelected[j] = d.strftime("%d/%m/%Y")
                rowSelected[j] =d
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected
#print(copyRange(copyList, copySheet))
#Paste range

#Paste data from copyRange into template sheet
def pasteRange(copyDict, pasteDict, sheetReceiving, copiedData):
    countRow = 0
    for i in range(12,14,1):
        for j in copyDict:
            sheetReceiving.cell(i,pasteDict[j] ).value = copiedData[countRow][j]
            
        countRow += 1

def createData():
    print("Processing...")
    selectedRange = copyRange(copyDict, copySheet) #Change the 4 number values
    pasteRange(copyDict, pasteDict, pasteSheet, selectedRange) #Change the 4 number values
    #You can save the template as another file to create a new file here too.s
    template.save("Sample2.xlsx")
    print("items copied and pasted!")
    return selectedRange

print(dateCopy)

copyDict= dictValidator(nonFilterdedCopyDict)
pasteDict= dictValidator(nonFilteredPasteDict)
print(createData())

#open all the files in a folder
# x= input("enter the path:")
# print(x)
# os.chdir(x)
# f= os.listdir(x)
# for i in f:
#     if i.endswith('.py') or i.endswith('.xlsx'):
#         print(i)
