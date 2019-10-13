#! Python 3
# - Copy and Paste Ranges using OpenPyXl library
import openpyxl ,datetime, os, xlrd
# excel loading
# file name is given, the output is sheets to copy from

def TheSheets(file):
    if file.endswith(".xls"):
        wb= xlrd.open_workbook(file)
        sheets = [ws for ws in wb.sheet_names() if ws =="EMC" or "AREA" in ws.upper()]
        return wb, sheets
    elif file.endswith(".xlsx"):
        wb= openpyxl.load_workbook(file)
        sheets = [ws for ws in wb.sheetnames if ws =="EMC" or "AREA" in ws.upper()]
        return wb, sheets

#Copy range of cells as a nested list.[{"Tag":"02-FT-010","Problem":"Blockage"},{"Tag":"04-FV-002","Problem":"Stuck"}]
def copyRange(copyDict, sheet):
    startRow = searchRowStarting(sheet, "TAG")[0]
    rangeSelected = []
    #Loops through selected Rows. the while is to loop until data is finished
    try:
        while "-" in str( sheet.cell(startRow, copyDict["Tag"]).value):
            rowSelected = {}
            for j in copyDict:
                rowSelected[j] =str(sheet.cell(startRow, copyDict[j]).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
            startRow= startRow + 1
    except IndexError:
            pass
    return rangeSelected

#Paste data from copyRange into template sheet
def pasteRange(copyDict, pasteDict, sheetReceiving, copiedData, datePaste):
    countRow = 0
    startRow= sheetReceiving.max_row +1
    #Check last row that it is not empty
    while str(sheetReceiving.cell(startRow-1, pasteDict["Tag"] ).value)== "None":
        startRow= startRow-1         #decrement the row to check if the data was manualy deleted

    endRow= startRow+ len(copiedData)
    # pasting is in here: #first we set the serial then the date
    for i in range(startRow, endRow,1):  # for every row, we start to paste the new row
        try:  #pasting a new serial number
            sheetReceiving.cell(i,1).value = sheetReceiving.cell(i-1,1).value +1
        except:
            sheetReceiving.cell(i,1).value= 1
        sheetReceiving.cell(i,pasteDict["Date"]).value = datePaste    # pasting date
        #  Second we paste corrsponding data: tag of copy with tag of paste, problem with problem, etc..
        for j in copyDict:   #for every column, we paste matched tags
            sheetReceiving.cell(i,pasteDict[j] ).value = copiedData[countRow][j]
        countRow += 1
#this returns the column place for the main headers ex. tag in column 2 and problem in column 3
def searchRowStarting(sheet, theWord):
    for i in range(1, 10,1):
        #Appends the row to a RowSelected list
        for j in range(1, 10,1):
            try:
                if theWord in str(sheet.cell(i,j).value).upper():
                    if '-' in str(sheet.cell(i+1,j).value):
                        return i+1, j, "first", sheet.cell(i+1,j).value
                    elif '-' in str(sheet.cell(i+2,j).value):
                        return i+2, j, "second", sheet.cell(i+2,j).value               
                    else:
                        return FileExistsError
            except IndexError:
                return FileExistsError

def searchForWord(sheet, theWord):
    for i in range(1, 10,1):
        #Appends the row to a RowSelected list
        for j in range(1, 10,1):
            if theWord in str(sheet.cell(i,j).value).upper():
                return i, j

def moveData(copyFileName, copyFileSheet, pasteFileName, pasteFileSheet):
    print("Processing...")

    if copyFileName.endswith(".xlsx"):
        wb= openpyxl.load_workbook(copyFileName)
        copySheet= wb[copyFileSheet]
    elif copyFileName.endswith(".xls"):
        wb = xlrd.open_workbook(copyFileName)
        copySheet= wb.sheet_by_name(copyFileSheet)

    pasteFile = openpyxl.load_workbook(pasteFileName)
    pasteSheet= pasteFile[pasteFileSheet]

    print("Files are successfully loaded!")
    pasteDict, copyDict, dateCopy= dictionaries( copyFileName, copySheet, pasteSheet)
    if searchRowStarting(copySheet,"TAG") != FileExistsError and searchRowStarting(copySheet,"TAG") != IndexError:
        selectedRange = copyRange(copyDict, copySheet)
        pasteRange(copyDict, pasteDict, pasteSheet, selectedRange, dateCopy) 
        #You can save the template as another file to create a new file here too
        pasteFile.save("D:/programs/python/excel/project/New folder/Sample2.xlsx")

        print("items copied and pasted!")
        return selectedRange, pasteFile

def dictionaries(copyFileName, copySheet, pasteSheet ):
    copyDict={}
    pasteDict= {}
    dateCopy = os.path.basename(copyFileName).split('.')[0]
    if "," in dateCopy:
        dateCopy= copyFileName.split(",")[0] +"-"+ copyFileName.split("-",maxsplit=1)[1]  
    dateCopy= str(dateCopy)

    try:
        pasteDict["Tag"]= searchForWord(pasteSheet, "TAG")[1]
    except TypeError:
        pass

    try:
        copyDict["Tag"]= searchForWord(copySheet, "TAG")[1]
    except TypeError:
        pass
    try:
        pasteDict["Problem"]= searchForWord(pasteSheet,"PROB")[1]
    except TypeError:
        pass
    try:
        copyDict["Problem"]= searchForWord(copySheet,"PROB")[1]
    except TypeError:
        pass
    try:
        pasteDict["Complain"]= searchForWord(pasteSheet,"COMPLAIN")[1]
    except TypeError:
        pass
    try:
        copyDict["Complain"]= searchForWord(copySheet,"COMPLAIN")[1]
    except:
        pass
    try:
        pasteDict["Action"]= searchForWord(pasteSheet,"ACTION")[1]
    except TypeError:
        pass
    try:
        copyDict["Action"]= searchForWord(copySheet,"ACTION")[1]
    except TypeError:
        pass
    try:
        pasteDict["Status"]= searchForWord(pasteSheet,"STATUS")[1]
    except TypeError:
        pass
    try:
        copyDict["Status"]= searchForWord(copySheet,"STATUS")[1]
    except TypeError:
        pass
    try:
        pasteDict["Date"]= searchForWord(pasteSheet,"DATE")[1]
    except TypeError:
        pass
    return pasteDict, copyDict, dateCopy