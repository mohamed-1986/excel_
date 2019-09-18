#! Python 3
# - Copy and Paste Ranges using OpenPyXl library

import os
import openpyxl
from pathlib import Path,PurePath
#Prepare the spreadsheets to copy from and paste too.

#File to be copied
# wb = openpyxl.load_workbook("Sample1.xlsx") #Add file name
# sheet = wb["Area 02"] #Add Sheet name

# #File to be pasted into
# template = openpyxl.load_workbook("Sample2.xlsx") #Add file name
# temp_sheet = template["Area 02"] #Add Sheet name

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected
         

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
def createData():
    print("Processing...")
    selectedRange = copyRange(1,8,5,11,sheet) #Change the 4 number values
    pastingRange = pasteRange(1,8,5,11,temp_sheet,selectedRange) #Change the 4 number values
    #You can save the template as another file to create a new file here too.s
    template.save("Sample2.xlsx")
    print("Range copied and pasted!")

#createData()
x= input("enter the path:")
#os.chdir(x)
# f= os.listdir(x)
# for i in f:
#     if i.endswith('.xlsx'):
#         print(i)

#to open all folders in the current path, here x = folder name only without the whole path
# PurePath add the complementary path:
# p = Path('.')
# g=[x for x in p.iterdir() if x.is_dir()]
# print(g)
# for eachFolder in g:
#     print(PurePath(os.getcwd(), eachFolder))

#to open all folders in the current path, here x = folder name with the whole path
p= Path(x)
allFolders=[y for y in p.iterdir() if y.is_dir()]

for eachFolder in allFolders:
    os.chdir(eachFolder)
    print(os.getcwd())
    f= os.listdir(eachFolder)
    for i in f:
        if i.endswith('.xls') or i.endswith('.xlsx'):
            wb = openpyxl.load_workbook(i) #Add file name
            sheet = wb["Area 02"] #Add Sheet name
            selectedRange = copyRange(1,8,5,11,sheet) #Change the 4 number values
            pastingRange = pasteRange(1,8,5,11,temp_sheet,selectedRange)